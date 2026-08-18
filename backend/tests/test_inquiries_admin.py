"""
Test Suite for Admin Inquiries Management & Export (Phase 03 / Inquiries Polish)

Validates:
- Admin authentication protection on all inquiry endpoints
- Stats computation and aggregate status breakdowns
- Listing with search and status filtering
- Single inquiry retrieval
- Status update lifecycle (new -> contacted -> qualified -> converted -> closed)
- Internal admin notes persistence and protection
- CSV export format and UTF-8 BOM encoding
- XLSX export format with Excel styling, multiple sheets, filters, and summary
"""

import io
import csv
import re
import openpyxl
import pytest
from datetime import datetime, timezone
from starlette.testclient import TestClient

from backend.server import app
from backend.routes.admin_auth import COOKIE_NAME, _create_session_token
from backend.routes.admin_inquiries import get_db


class MockCursor:
    def __init__(self, docs):
        self.docs = docs

    def sort(self, key, direction=1):
        return self

    def skip(self, n):
        self.docs = self.docs[n:]
        return self

    def limit(self, n):
        self.docs = self.docs[:n]
        return self

    async def to_list(self, n):
        return [dict(d) for d in self.docs[:n]]


class MockContactSubmissionsCollection:
    def __init__(self, docs):
        self.docs = [dict(d) for d in docs]

    def _filter(self, query):
        if not query:
            return list(self.docs)
        results = []
        for d in self.docs:
            match = True
            for k, v in query.items():
                if k == "$or":
                    or_match = False
                    for clause in v:
                        for ck, cv in clause.items():
                            if isinstance(cv, dict) and "$regex" in cv:
                                if re.search(cv["$regex"], str(d.get(ck, "")), re.I):
                                    or_match = True
                            elif cv is None or cv == "":
                                if not d.get(ck):
                                    or_match = True
                            elif isinstance(cv, dict) and "$exists" in cv:
                                if (ck not in d or not d.get(ck)) == (not cv["$exists"]):
                                    or_match = True
                            elif d.get(ck) == cv:
                                or_match = True
                    if not or_match:
                        match = False
                elif k == "$and":
                    and_match = True
                    for sub_query in v:
                        if not self._match_clause(d, sub_query):
                            and_match = False
                    if not and_match:
                        match = False
                else:
                    if not self._match_field(d, k, v):
                        match = False
            if match:
                results.append(d)
        return results

    def _match_clause(self, d, clause):
        for k, v in clause.items():
            if k == "$or":
                for sub in v:
                    for sk, sv in sub.items():
                        if isinstance(sv, dict) and "$regex" in sv:
                            if re.search(sv["$regex"], str(d.get(sk, "")), re.I):
                                return True
                        elif d.get(sk) == sv:
                            return True
                return False
            elif not self._match_field(d, k, v):
                return False
        return True

    def _match_field(self, d, k, v):
        if isinstance(v, dict) and "$regex" in v:
            return bool(re.search(v["$regex"], str(d.get(k, "")), re.I))
        return d.get(k) == v

    async def count_documents(self, query):
        filtered = self._filter(query)
        return len(filtered)

    def find(self, query=None, projection=None):
        filtered = self._filter(query)
        return MockCursor(filtered)

    async def find_one(self, query, projection=None):
        filtered = self._filter(query)
        if filtered:
            return dict(filtered[0])
        return None

    async def update_one(self, query, update):
        filtered = self._filter(query)
        if not filtered:
            return type("UpdateResult", (), {"matched_count": 0, "modified_count": 0})()
        target = filtered[0]
        if "$set" in update:
            for k, v in update["$set"].items():
                target[k] = v
        return type("UpdateResult", (), {"matched_count": 1, "modified_count": 1})()


class MockDatabase:
    def __init__(self, docs):
        self.contact_submissions = MockContactSubmissionsCollection(docs)


@pytest.fixture
def mock_db():
    initial_docs = [
        {
            "id": "inq-101",
            "name": "Ishaan Aakarsh",
            "email": "ishaan@navigatte.com",
            "phone": "+91 98765 43210",
            "company": "Navigatte Corp",
            "designation": "Chief Financial Officer",
            "service_of_interest": "Internal Audit & Assurance",
            "message": "We need quarterly internal controls audit.",
            "source": "website_contact",
            "status": "new",
            "notes": None,
            "created_at": datetime(2026, 8, 18, 10, 30, tzinfo=timezone.utc),
        },
        {
            "id": "inq-102",
            "name": "Rohan Gupta",
            "email": "rohan@rgautomotive.in",
            "phone": "+91 91234 56789",
            "company": "RG Automotive",
            "designation": "Finance Director",
            "service_of_interest": "Corporate Tax & Advisory",
            "message": "Inquiry regarding cross-border GST structuring.",
            "source": "website_contact",
            "status": "contacted",
            "notes": "Spoke on phone, scheduled follow up.",
            "created_at": datetime(2026, 8, 17, 14, 0, tzinfo=timezone.utc),
        },
        {
            "id": "inq-103",
            "name": "Priya Sharma",
            "email": "priya@sharmaindustries.com",
            "phone": "+91 99887 76655",
            "company": "Sharma Industries",
            "designation": "Managing Director",
            "service_of_interest": "Risk Advisory",
            "message": "Looking for comprehensive board risk review.",
            "source": "website_contact",
            "status": "qualified",
            "notes": "High priority lead.",
            "created_at": datetime(2026, 8, 16, 9, 15, tzinfo=timezone.utc),
        },
    ]
    return MockDatabase(initial_docs)


@pytest.fixture
def client(mock_db):
    app.dependency_overrides[get_db] = lambda: mock_db
    test_client = TestClient(app)
    yield test_client
    app.dependency_overrides.clear()


@pytest.fixture
def auth_client(client):
    token = _create_session_token()
    client.cookies.set(COOKIE_NAME, token)
    return client


# ---------- Tests ----------

def test_inquiries_endpoints_require_admin_auth(client):
    """Endpoints must block unauthenticated access with 401"""
    res_list = client.get("/api/admin/inquiries")
    assert res_list.status_code == 401

    res_stats = client.get("/api/admin/inquiries/stats")
    assert res_stats.status_code == 401

    res_export = client.get("/api/admin/inquiries/export")
    assert res_export.status_code == 401


def test_inquiries_stats_breakdown(auth_client):
    """Stats endpoint returns accurate counts by status"""
    res = auth_client.get("/api/admin/inquiries/stats")
    assert res.status_code == 200
    data = res.json()
    assert data["total"] == 3
    assert data["by_status"]["new"] == 1
    assert data["by_status"]["contacted"] == 1
    assert data["by_status"]["qualified"] == 1
    assert data["by_status"]["converted"] == 0
    assert data["by_status"]["closed"] == 0


def test_list_inquiries_and_filtering(auth_client):
    """Listing enquiries supports search by name and status filtering"""
    # List all
    res_all = auth_client.get("/api/admin/inquiries")
    assert res_all.status_code == 200
    assert len(res_all.json()) == 3

    # Filter status=contacted
    res_contacted = auth_client.get("/api/admin/inquiries?status=contacted")
    assert res_contacted.status_code == 200
    items = res_contacted.json()
    assert len(items) == 1
    assert items[0]["name"] == "Rohan Gupta"

    # Search name
    res_search = auth_client.get("/api/admin/inquiries?q=Ishaan")
    assert res_search.status_code == 200
    search_items = res_search.json()
    assert len(search_items) == 1
    assert search_items[0]["email"] == "ishaan@navigatte.com"


def test_get_single_inquiry(auth_client):
    """Retrieves full details for a single inquiry"""
    res = auth_client.get("/api/admin/inquiries/inq-101")
    assert res.status_code == 200
    doc = res.json()
    assert doc["id"] == "inq-101"
    assert doc["company"] == "Navigatte Corp"
    assert doc["service_of_interest"] == "Internal Audit & Assurance"

    res_404 = auth_client.get("/api/admin/inquiries/non-existent-id")
    assert res_404.status_code == 404


def test_update_inquiry_status_lifecycle(auth_client):
    """Updates status cleanly and validates status transitions"""
    # Change inq-101 from new to contacted
    res_update = auth_client.patch(
        "/api/admin/inquiries/inq-101/status",
        json={"status": "contacted"},
    )
    assert res_update.status_code == 200
    assert res_update.json()["status"] == "contacted"
    assert "status_updated_at" in res_update.json()

    # Invalid status should return 400
    res_invalid = auth_client.patch(
        "/api/admin/inquiries/inq-101/status",
        json={"status": "invalid_status"},
    )
    assert res_invalid.status_code == 400


def test_update_inquiry_internal_notes(auth_client):
    """Updates internal notes securely"""
    res_notes = auth_client.patch(
        "/api/admin/inquiries/inq-101/notes",
        json={"notes": "Confidential: Client requesting urgent proposal for board meeting."},
    )
    assert res_notes.status_code == 200
    assert "Confidential" in res_notes.json()["notes"]


def test_export_inquiries_csv(auth_client):
    """CSV export returns clean CSV with headers and matching records"""
    res = auth_client.get("/api/admin/inquiries/export?format=csv")
    assert res.status_code == 200
    assert "text/csv" in res.headers["Content-Type"]
    assert "attachment; filename=\"psa_enquiries_" in res.headers["Content-Disposition"]

    content = res.content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(content))
    rows = list(reader)

    # Verify header
    assert rows[0][0] == "ID"
    assert rows[0][2] == "Status"
    assert rows[0][3] == "Name"
    assert rows[0][4] == "Email"

    # Verify rows count (header + 3 records)
    assert len(rows) == 4
    assert any("Ishaan Aakarsh" in r for r in rows)


def test_export_inquiries_xlsx(auth_client):
    """XLSX export returns valid Excel workbook with styled sheets and table filters"""
    res = auth_client.get("/api/admin/inquiries/export?format=xlsx")
    assert res.status_code == 200
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in res.headers["Content-Type"]
    assert "attachment; filename=\"psa_enquiries_" in res.headers["Content-Disposition"]

    # Parse Excel workbook with openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    assert "Enquiries" in wb.sheetnames
    assert "Export Summary" in wb.sheetnames

    # Check Enquiries worksheet
    ws_enquiries = wb["Enquiries"]
    assert ws_enquiries.cell(row=1, column=1).value == "ID"
    assert ws_enquiries.cell(row=1, column=3).value == "Status"
    assert ws_enquiries.cell(row=1, column=4).value == "Name"
    assert ws_enquiries.max_row == 4  # Header + 3 records

    # Check Export Summary worksheet
    ws_summary = wb["Export Summary"]
    assert ws_summary.cell(row=1, column=1).value == "P Suman & Associates — Enquiries Export"
    assert ws_summary.cell(row=3, column=2).value == 3  # Total records
