"""
Admin Inquiries Router — Phase 03

Provides authenticated admin endpoints for managing website-generated enquiries
stored in the contact_submissions collection.

Inquiry Status Workflow:
    new → contacted → qualified → converted → closed

Source taxonomy (V1, extensible):
    website_contact  — Contact page form (Phase 01)
    (future: calcom_meeting, newsletter_lead, etc.)

All endpoints require a valid admin session cookie via get_current_admin.
"""

import io
import csv
import logging
import re
from datetime import datetime, timezone
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorDatabase
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.core.auth import get_current_admin

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/admin/inquiries", tags=["Admin Inquiries"])

VALID_STATUSES = {"new", "contacted", "qualified", "converted", "closed"}


# ---------- DB dependency ----------

def get_db() -> AsyncIOMotorDatabase:
    from backend.server import db
    return db


# ---------- Pydantic models ----------

class InquiryStatusUpdate(BaseModel):
    status: str


class InquiryNotesUpdate(BaseModel):
    notes: Optional[str] = None


# ---------- Helpers ----------

def _normalise_doc(doc: dict) -> dict:
    """
    Ensures all required fields have sensible defaults for documents
    created before Phase 03 (which lacked source/status/notes).
    """
    doc.pop("_id", None)
    # Backward compat defaults
    if not doc.get("source"):
        doc["source"] = "website_contact"
    if not doc.get("status"):
        doc["status"] = "new"
    # Normalise created_at to ISO string for consistent serialisation
    ca = doc.get("created_at")
    if isinstance(ca, datetime):
        doc["created_at"] = ca.isoformat()
    return doc


def _build_filter_query(
    q: Optional[str] = None,
    status: Optional[str] = None,
    source: Optional[str] = None,
) -> dict:
    filter_query: dict = {}

    if status:
        if status not in VALID_STATUSES:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid status. Valid: {sorted(VALID_STATUSES)}",
            )
        if status == "new":
            # Include legacy docs that have no status field
            filter_query["$or"] = [
                {"status": "new"},
                {"status": {"$exists": False}},
                {"status": None},
                {"status": ""},
            ]
        else:
            filter_query["status"] = status

    if source:
        filter_query["source"] = source

    if q and q.strip():
        term = q.strip()
        regex = {"$regex": re.escape(term), "$options": "i"}
        search_clause = {"$or": [{"name": regex}, {"email": regex}, {"company": regex}]}
        if "$or" in filter_query:
            # Combine existing $or (status) with search via $and
            filter_query = {"$and": [filter_query, search_clause]}
        else:
            filter_query.update(search_clause)

    return filter_query


# ---------- Endpoints ----------

@router.get("/stats", dependencies=[Depends(get_current_admin)])
async def get_inquiry_stats(db: AsyncIOMotorDatabase = Depends(get_db)):
    """
    Returns total count and per-status breakdown of all enquiries.
    Treats missing/null status as 'new' for aggregate purposes.
    """
    total = await db.contact_submissions.count_documents({})

    # Count per explicit status value
    status_counts = {}
    for s in VALID_STATUSES:
        status_counts[s] = await db.contact_submissions.count_documents({"status": s})

    # Documents with no status field (legacy) → counted as 'new'
    no_status = await db.contact_submissions.count_documents(
        {"$or": [{"status": {"$exists": False}}, {"status": None}, {"status": ""}]}
    )
    status_counts["new"] = status_counts.get("new", 0) + no_status

    return {
        "total": total,
        "by_status": status_counts,
    }


@router.get("/export", dependencies=[Depends(get_current_admin)])
async def export_inquiries(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    q: Optional[str] = Query(None, description="Search term"),
    status: Optional[str] = Query(None, description="Status filter"),
    source: Optional[str] = Query(None, description="Source filter"),
    db: AsyncIOMotorDatabase = Depends(get_db),
):
    """
    Exports enquiries to a formatted XLSX spreadsheet or standard CSV file.
    Supports filtering by status, source, or search query.
    Admin authentication is strictly required.
    """
    filter_query = _build_filter_query(q=q, status=status, source=source)
    raw_docs = (
        await db.contact_submissions.find(filter_query, {"_id": 0})
        .sort("created_at", -1)
        .to_list(10000)
    )
    docs = [_normalise_doc(d) for d in raw_docs]

    timestamp_str = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)
        
        writer.writerow([
            "ID",
            "Received Date (UTC)",
            "Status",
            "Name",
            "Email",
            "Phone",
            "Company",
            "Designation",
            "Service of Interest",
            "Source",
            "Message",
            "Internal Notes",
            "Last Updated",
        ])

        for d in docs:
            writer.writerow([
                d.get("id", ""),
                d.get("created_at", ""),
                d.get("status", "new").capitalize(),
                d.get("name", ""),
                d.get("email", ""),
                d.get("phone", "") or "",
                d.get("company", "") or "",
                d.get("designation", "") or "",
                d.get("service_of_interest", "") or "General Inquiry",
                d.get("source", "website_contact"),
                d.get("message", ""),
                d.get("notes", "") or "",
                d.get("status_updated_at", "") or "",
            ])

        csv_bytes = output.getvalue().encode("utf-8-sig")
        filename = f"psa_enquiries_{timestamp_str}.csv"
        return StreamingResponse(
            io.BytesIO(csv_bytes),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    # XLSX Generation
    wb = openpyxl.Workbook()

    # Worksheet 1: Enquiries
    ws = wb.active
    ws.title = "Enquiries"
    ws.views.sheetView[0].showGridLines = True

    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="0A2540", end_color="0A2540", fill_type="solid")
    align_header = Alignment(horizontal="center", vertical="center", wrap_text=False)

    font_data = Font(name="Segoe UI", size=10, color="0F172A")
    fill_even = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    border_thin = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    headers = [
        ("ID", 32),
        ("Received Date (UTC)", 20),
        ("Status", 14),
        ("Name", 24),
        ("Email", 28),
        ("Phone", 18),
        ("Company", 24),
        ("Designation", 20),
        ("Service of Interest", 26),
        ("Source", 18),
        ("Message", 45),
        ("Internal Notes", 35),
        ("Last Updated", 20),
    ]

    ws.row_dimensions[1].height = 28
    for col_idx, (header_title, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header_title)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_thin
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    for row_idx, d in enumerate(docs, start=2):
        ws.row_dimensions[row_idx].height = 24
        fill_row = fill_even if row_idx % 2 == 0 else fill_odd

        row_values = [
            d.get("id", ""),
            d.get("created_at", ""),
            d.get("status", "new").capitalize(),
            d.get("name", ""),
            d.get("email", ""),
            d.get("phone", "") or "—",
            d.get("company", "") or "—",
            d.get("designation", "") or "—",
            d.get("service_of_interest", "") or "General Inquiry",
            d.get("source", "website_contact"),
            d.get("message", ""),
            d.get("notes", "") or "—",
            d.get("status_updated_at", "") or "—",
        ]

        for col_idx, val in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_data
            cell.fill = fill_row
            cell.border = border_thin
            
            if col_idx in (1, 2, 3, 13):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in (11, 12):
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = "A2"
    if docs:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(docs) + 1}"

    # Worksheet 2: Export Summary
    ws_summary = wb.create_sheet(title="Export Summary")
    ws_summary.views.sheetView[0].showGridLines = True
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 35

    summary_rows = [
        ("P Suman & Associates — Enquiries Export", ""),
        ("Export Generated (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")),
        ("Total Records Exported", len(docs)),
        ("Filter Status", status.capitalize() if status else "All Statuses"),
        ("Filter Search Query", q if q else "(None)"),
        ("Filter Source", source if source else "(All Sources)"),
        ("", ""),
        ("Breakdown by Status", ""),
    ]

    status_counts = {}
    for d in docs:
        st = d.get("status", "new").capitalize()
        status_counts[st] = status_counts.get(st, 0) + 1
    
    for st_name in ["New", "Contacted", "Qualified", "Converted", "Closed"]:
        summary_rows.append((f"  • {st_name}", status_counts.get(st_name, 0)))

    ws_summary.row_dimensions[1].height = 26
    for row_idx, (k, v) in enumerate(summary_rows, 1):
        cell_a = ws_summary.cell(row=row_idx, column=1, value=k)
        cell_b = ws_summary.cell(row=row_idx, column=2, value=v)
        
        if row_idx == 1:
            cell_a.font = Font(name="Segoe UI", size=12, bold=True, color="0A2540")
        elif k.startswith("Breakdown"):
            cell_a.font = Font(name="Segoe UI", size=11, bold=True, color="0A2540")
        else:
            cell_a.font = Font(name="Segoe UI", size=10, bold=True, color="475569")
            cell_b.font = Font(name="Segoe UI", size=10, color="0F172A")

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"psa_enquiries_{timestamp_str}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("", dependencies=[Depends(get_current_admin)])
async def list_inquiries(
    q: Optional[str] = Query(None, description="Search name, email, or company"),
    status: Optional[str] = Query(None, description="Filter by status"),
    source: Optional[str] = Query(None, description="Filter by source"),
    limit: int = Query(50, le=200),
    skip: int = Query(0),
    db: AsyncIOMotorDatabase = Depends(get_db),
):
    """
    List enquiries with optional search and filter.
    Documents lacking a status field are treated as 'new'.
    """
    filter_query = _build_filter_query(q=q, status=status, source=source)

    docs = (
        await db.contact_submissions.find(filter_query, {"_id": 0})
        .sort("created_at", -1)
        .skip(skip)
        .limit(limit)
        .to_list(limit)
    )

    return [_normalise_doc(d) for d in docs]



@router.get("/{inquiry_id}", dependencies=[Depends(get_current_admin)])
async def get_inquiry(inquiry_id: str, db: AsyncIOMotorDatabase = Depends(get_db)):
    """
    Retrieve a single enquiry by its id field.
    """
    doc = await db.contact_submissions.find_one({"id": inquiry_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Inquiry not found")
    return _normalise_doc(doc)


@router.patch("/{inquiry_id}/status", dependencies=[Depends(get_current_admin)])
async def update_inquiry_status(
    inquiry_id: str,
    payload: InquiryStatusUpdate,
    db: AsyncIOMotorDatabase = Depends(get_db),
):
    """
    Update the status of a single enquiry. Persists immediately.
    """
    new_status = payload.status.strip().lower()
    if new_status not in VALID_STATUSES:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid status '{new_status}'. Valid values: {sorted(VALID_STATUSES)}",
        )

    result = await db.contact_submissions.update_one(
        {"id": inquiry_id},
        {"$set": {"status": new_status, "status_updated_at": datetime.now(timezone.utc).isoformat()}},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Inquiry not found")

    doc = await db.contact_submissions.find_one({"id": inquiry_id}, {"_id": 0})
    return _normalise_doc(doc)


@router.patch("/{inquiry_id}/notes", dependencies=[Depends(get_current_admin)])
async def update_inquiry_notes(
    inquiry_id: str,
    payload: InquiryNotesUpdate,
    db: AsyncIOMotorDatabase = Depends(get_db),
):
    """
    Update admin notes on an enquiry. Notes are internal only.
    """
    result = await db.contact_submissions.update_one(
        {"id": inquiry_id},
        {"$set": {"notes": payload.notes}},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Inquiry not found")

    doc = await db.contact_submissions.find_one({"id": inquiry_id}, {"_id": 0})
    return _normalise_doc(doc)
